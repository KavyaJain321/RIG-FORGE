from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = [
    ("Pragati Chamoli",  "pragati@rigforge.com",  "2B4BgPc@HK%C"),
    ("Shakshi Verma",    "shakshi@rigforge.com",  "E&LLaDa2h$8u"),
    ("Minakshi Uniyal",  "minakshi@rigforge.com", "#rLqjwB@7mys"),
    ("Tanisha Sharma",   "tanisha@rigforge.com",  "DD&!F3LmGQ7q"),
    ("Armaan Juneja",    "armaan@rigforge.com",   "VNAaW%XcWp&Q"),
    ("Aashray Iyengar",  "aashray@rigforge.com",  "u9!Mm5q&rXPK"),
    ("Anamika Ghuman",   "anamika@rigforge.com",  "E&bCF2!kjf2u"),
    ("Karnika Karanwal", "karnika@rigforge.com",  "Skypna@#kzZs"),
    ("Aditi Bhidola",    "aditi@rigforge.com",    "NJ%@LHGpVsvw"),
    ("Rittana Mittal",   "rittana@rigforge.com",  "gMs#vh7peQ&h"),
]

LOGIN_URL = "https://rig-forge.onrender.com/login"

wb = Workbook()
ws = wb.active
ws.title = "News Prism Onboarding"

ARIAL = "Arial"
MONO = "Consolas"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "RIG-FORGE  —  News Prism · Intern Credentials"
ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", start_color="1F2937")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 32

ws["A2"] = (
    "Share each row privately. Every account below has 'Must Change Password' enabled — "
    "on first login the application will prompt them to set a new password. "
    "All 10 are added as EMPLOYEE members of the News Prism project."
)
ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="4B5563")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 38

headers = ["#", "Name", "Email", "Role", "Temporary Password", "Login URL"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="374151")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws.row_dimensions[4].height = 24

for i, (name, email, pwd) in enumerate(rows, start=1):
    r = 4 + i
    ws.cell(row=r, column=1, value=i).font = Font(name=ARIAL, size=10, color="6B7280")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=2, value=name).font = Font(name=ARIAL, size=10)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=r, column=3, value=email).font = Font(name=ARIAL, size=10)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=r, column=4, value="EMPLOYEE").font = Font(name=ARIAL, size=10, bold=True, color="166534")
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=5, value=pwd).font = Font(name=MONO, size=11, bold=True, color="B91C1C")
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=r, column=6, value=LOGIN_URL).font = Font(name=ARIAL, size=10, color="2563EB", underline="single")
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=r, column=6).hyperlink = LOGIN_URL
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = border
    if i % 2 == 0:
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color="F9FAFB")

note_r = 4 + len(rows) + 2
ws.cell(row=note_r, column=1, value="Note:  All 10 users are members of the 'News Prism' project. Password change is required on first login.")
ws.cell(row=note_r, column=1).font = Font(name=ARIAL, size=10, italic=True, color="4B5563")
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=6)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 38

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

out = "C:/Users/Jain/Desktop/RIG-FORGE-main/RIG-FORGE-main/news-prism-onboarding.xlsx"
wb.save(out)
print(f"wrote: {out}  ({len(rows)} rows)")
