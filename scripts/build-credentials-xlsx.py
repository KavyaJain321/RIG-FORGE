from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = [
    ("Pranav",   "pranavv@rigforge.com",  "ADMIN",    "qsa$qmpQ$amJ"),
    ("Abhyam",   "abhyam@rigforge.com",   "EMPLOYEE", "SEG9Y$Phf$J7"),
    ("Ahmed",    "ahmed@rigforge.com",    "EMPLOYEE", "P&CcsFNQbzM5"),
    ("Daksh",    "daksh@rigforge.com",    "EMPLOYEE", "RyfBeXXC3&ey"),
    ("Kashvi",   "kashvi@rigforge.com",   "EMPLOYEE", "PJg$A5HzgUGb"),
    ("Krishn",   "krishn@rigforge.com",   "EMPLOYEE", "!ns3z%Zm3QtW"),
    ("Pankaj",   "pankaj@rigforge.com",   "EMPLOYEE", "hmH%P2Vyk3eh"),
    ("Radhesh",  "rhadesh@rigforge.com",  "EMPLOYEE", "zS&75G4S#uY7"),
    ("Rohan",    "rohun@rigforge.com",    "EMPLOYEE", "2&6vsjbv7$ZF"),
    ("Shubham",  "shubham@rigforge.com",  "EMPLOYEE", "zPGg7m4%EAj2"),
    ("Sudipta",  "sudipta@rigforge.com",  "EMPLOYEE", "jvHA85apZgX#"),
    ("Sumit",    "sumit@rigforge.com",    "EMPLOYEE", "8L6DFaS3@t5K"),
    ("Utkarsh",  "utkarsh@rigforge.com",  "EMPLOYEE", "kTQ8VG7D@7Bv"),
    ("Yash",     "yash@rigforge.com",     "EMPLOYEE", "@2DMFUubFrdp"),
]

LOGIN_URL = "http://localhost:3000/login"
NOTE = "You must change this password on first login."

wb = Workbook()
ws = wb.active
ws.title = "Credentials"

ARIAL = "Arial"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "RIG-FORGE  —  Employee Credentials"
ws["A1"].font = Font(name=ARIAL, size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", start_color="1F2937")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 32

ws["A2"] = (
    "Share the row corresponding to each person with them privately. "
    "Every account below has `Must Change Password` enabled — on first login "
    "the application will prompt them to set a new password."
)
ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="4B5563")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 32

headers = ["#", "Name", "Email", "Role", "Temporary Password", "Login URL"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="374151")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws.row_dimensions[4].height = 24

MONO = "Consolas"
for i, (name, email, role, pwd) in enumerate(rows, start=1):
    r = 4 + i
    values = [i, name, email, role, pwd, LOGIN_URL]
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.border = border
        if col == 5:
            c.font = Font(name=MONO, size=11, bold=True, color="B91C1C")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        elif col == 4:
            c.font = Font(name=ARIAL, size=10, bold=True,
                          color="1D4ED8" if role == "ADMIN" else "166534")
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 1:
            c.font = Font(name=ARIAL, size=10, color="6B7280")
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 6:
            c.font = Font(name=ARIAL, size=10, color="2563EB", underline="single")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.hyperlink = LOGIN_URL
        else:
            c.font = Font(name=ARIAL, size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if i % 2 == 0:
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color="F9FAFB")

note_r = 4 + len(rows) + 2
ws.cell(row=note_r, column=1, value=f"Note:  {NOTE}")
ws.cell(row=note_r, column=1).font = Font(name=ARIAL, size=10, italic=True, color="4B5563")
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=6)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 34

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

out = "C:/Users/Jain/Desktop/RIG-FORGE-main/RIG-FORGE-main/employee_credentials.xlsx"
wb.save(out)
print(f"wrote: {out}  ({len(rows)} rows)")
