from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()
ws = wb.active
ws.title = "Credentials"

# ── Colour palette ────────────────────────────────────────────────────────────
C_DARK_BG    = "0D0D0D"
C_ACCENT     = "00FF88"       # RIG-FORGE green accent
C_SUPER      = "FFD700"       # gold  – SUPER_ADMIN
C_ADMIN      = "4FC3F7"       # sky   – ADMIN
C_EMP        = "A5D6A7"       # mint  – EMPLOYEE
C_HEADER_FG  = "FFFFFF"
C_WARN_FG    = "FF6B6B"
C_MUTED      = "888888"
C_BORDER     = "2A2A2A"
C_ROW_ALT    = "1A1A1A"
C_ROW_BASE   = "141414"

thin = Side(style="thin", color=C_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mono(size=9, bold=False, color="E0E0E0"):
    return Font(name="Courier New", size=size, bold=bold, color=color)

def sans(size=9, bold=False, color="E0E0E0"):
    return Font(name="Arial", size=size, bold=bold, color=color)

center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Title block (rows 1-4) ────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
ws["A1"] = "RIG FORGE — CONFIDENTIAL CREDENTIALS"
ws["A1"].font      = Font(name="Arial", size=16, bold=True, color=C_ACCENT)
ws["A1"].fill      = fill(C_DARK_BG)
ws["A1"].alignment = center

ws.merge_cells("A2:F2")
ws["A2"] = "WORKFORCE INTELLIGENCE PLATFORM · INTERNAL USE ONLY"
ws["A2"].font      = Font(name="Arial", size=9, bold=False, color=C_MUTED)
ws["A2"].fill      = fill(C_DARK_BG)
ws["A2"].alignment = center

ws.merge_cells("A3:F3")
ws["A3"] = f"Generated: {datetime.datetime.now().strftime('%d %B %Y, %H:%M')}"
ws["A3"].font      = Font(name="Arial", size=8, color=C_MUTED)
ws["A3"].fill      = fill(C_DARK_BG)
ws["A3"].alignment = center

ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 14
ws.row_dimensions[4].height = 8   # spacer

# ── Column headers (row 5) ────────────────────────────────────────────────────
headers = ["#", "NAME", "EMAIL ADDRESS", "ROLE", "TEMPORARY PASSWORD", "MUST CHANGE PASSWORD?"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=h)
    cell.font      = Font(name="Arial", size=9, bold=True, color=C_DARK_BG)
    cell.fill      = fill(C_ACCENT)
    cell.alignment = center
    cell.border    = border
ws.row_dimensions[5].height = 20

# ── Credential data ───────────────────────────────────────────────────────────
users = [
    ("Rohit Gandhi",  "rohit@rigforge.com",   "SUPER_ADMIN", "RigForge@Rohit#2024", "No"),
    ("Kavya Jain",    "kavya@rigforge.com",   "ADMIN",       "t8NHBZY$qhRx",        "Yes — on first login"),
    ("Pranav",        "pranavv@rigforge.com", "ADMIN",       "EvQ&dVS%5LPd",        "Yes — on first login"),
    ("Abhyam",        "abhyam@rigforge.com",  "EMPLOYEE",    "d6VfMYaG#Gz7",        "Yes — on first login"),
    ("Ahmed",         "ahmed@rigforge.com",   "EMPLOYEE",    "P&CcsFNQbzM5",        "Yes — on first login"),
    ("Daksh",         "daksh@rigforge.com",   "EMPLOYEE",    "RyfBeXXC3&ey",        "Yes — on first login"),
    ("Kashvi",        "kashvi@rigforge.com",  "EMPLOYEE",    "PJg$A5HzgUGb",        "Yes — on first login"),
    ("Krishn",        "krishn@rigforge.com",  "EMPLOYEE",    "!ns3z%Zm3QtW",        "Yes — on first login"),
    ("Pankaj",        "pankaj@rigforge.com",  "EMPLOYEE",    "hmH%P2Vyk3eh",        "Yes — on first login"),
    ("Radhesh",       "rhadesh@rigforge.com", "EMPLOYEE",    "z$3zr!AEwdB8",        "Yes — on first login"),
    ("Rohan",         "rohun@rigforge.com",   "EMPLOYEE",    "2&6vsjbv7$ZF",        "Yes — on first login"),
    ("Shubham",       "shubham@rigforge.com", "EMPLOYEE",    "zPGg7m4%EAj2",        "Yes — on first login"),
    ("Sudipta",       "sudipta@rigforge.com", "EMPLOYEE",    "jvHA85apZgX#",        "Yes — on first login"),
    ("Sumit",         "sumit@rigforge.com",   "EMPLOYEE",    "8L6DFaS3@t5K",        "Yes — on first login"),
    ("Utkarsh",       "utkarsh@rigforge.com", "EMPLOYEE",    "kTQ8VG7D@7Bv",        "Yes — on first login"),
    ("Yash",          "yash@rigforge.com",    "EMPLOYEE",    "@2DMFUubFrdp",        "Yes — on first login"),
]

ROLE_COLORS = {
    "SUPER_ADMIN": C_SUPER,
    "ADMIN":       C_ADMIN,
    "EMPLOYEE":    C_EMP,
}

for i, (name, email, role, pwd, must_change) in enumerate(users):
    row = 6 + i
    row_bg = C_ROW_ALT if i % 2 else C_ROW_BASE
    role_fg = ROLE_COLORS.get(role, "E0E0E0")

    data = [i + 1, name, email, role, pwd, must_change]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill   = fill(row_bg)
        cell.border = border

        if col == 1:   # index
            cell.font      = sans(8, color=C_MUTED)
            cell.alignment = center
        elif col == 2: # name
            cell.font      = sans(9, bold=True, color="F0F0F0")
            cell.alignment = left
        elif col == 3: # email
            cell.font      = mono(9, color="A0C8FF")
            cell.alignment = left
        elif col == 4: # role
            cell.font      = Font(name="Arial", size=8, bold=True, color=C_DARK_BG)
            cell.fill      = fill(role_fg)
            cell.alignment = center
        elif col == 5: # password
            cell.font      = mono(10, bold=True, color="FFEB3B")
            cell.alignment = center
        elif col == 6: # must change
            if must_change.startswith("Yes"):
                cell.font = sans(8, bold=True, color=C_WARN_FG)
            else:
                cell.font = sans(8, bold=True, color=C_ACCENT)
            cell.alignment = center

    ws.row_dimensions[row].height = 18

# ── Warning notice (after data) ───────────────────────────────────────────────
notice_row = 6 + len(users) + 1
ws.merge_cells(f"A{notice_row}:F{notice_row}")
ws[f"A{notice_row}"] = (
    "⚠  CONFIDENTIAL — Do not share this file publicly. "
    "All passwords marked 'Yes' must be changed on first login. "
    "Users who do not change their password will be locked out of the dashboard."
)
ws[f"A{notice_row}"].font      = Font(name="Arial", size=8, bold=True, color=C_WARN_FG)
ws[f"A{notice_row}"].fill      = fill("1E0000")
ws[f"A{notice_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws[f"A{notice_row}"].border    = Border(
    left=Side(style="medium", color=C_WARN_FG),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)
ws.row_dimensions[notice_row].height = 36

# ── Legend (after notice) ─────────────────────────────────────────────────────
legend_row = notice_row + 2
ws.merge_cells(f"A{legend_row}:F{legend_row}")
ws[f"A{legend_row}"] = "ROLE LEGEND"
ws[f"A{legend_row}"].font      = Font(name="Arial", size=8, bold=True, color=C_ACCENT)
ws[f"A{legend_row}"].fill      = fill(C_DARK_BG)
ws[f"A{legend_row}"].alignment = left

for idx, (role_label, role_color, desc) in enumerate([
    ("SUPER_ADMIN", C_SUPER,  "Rohit Gandhi — full system control, can create admin accounts"),
    ("ADMIN",       C_ADMIN,  "Tech leads / 2nd point of contact — manage projects & employees"),
    ("EMPLOYEE",    C_EMP,    "Team members — view assigned projects, log work, raise tickets"),
]):
    r = legend_row + 1 + idx
    ws.cell(r, 1, role_label).font      = Font(name="Arial", size=8, bold=True, color=C_DARK_BG)
    ws.cell(r, 1).fill      = fill(role_color)
    ws.cell(r, 1).alignment = center
    ws.cell(r, 1).border    = border
    ws.merge_cells(f"B{r}:F{r}")
    ws.cell(r, 2, desc).font      = sans(8, color="C0C0C0")
    ws.cell(r, 2).fill      = fill(C_ROW_BASE)
    ws.cell(r, 2).alignment = left
    ws.cell(r, 2).border    = border
    ws.row_dimensions[r].height = 16

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [4, 18, 28, 14, 22, 26]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze panes at row 6 ────────────────────────────────────────────────────
ws.freeze_panes = "A6"

# ── Sheet tab colour ──────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = C_ACCENT

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"C:\Users\Jain\Desktop\RigForge_Credentials.xlsx"
wb.save(out)
print(f"Saved: {out}")
